
import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
import requests
from collections import OrderedDict
import json
import pandas as pd
import openpyxl
import os
from datetime import date

def index(request):

    context = {}
    if request.method == "POST":

        if 'odbe_import' in request.POST:
            print('ODBE import gomb megnyomva')
            odbe_excel_import()

        if 'kvgomb' in request.POST:
            print('Készletváltozás gomb megnyomva')
            kvaltozas_import()

    return  render(request, 'trafik_app/index.html', context)

def kvaltozas_import():
    path = '/Users/futurex/Desktop/készletváltozás 02-20 02-26.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    last_row = sheet.max_row
    keszletvaltozas = []

    for x in range(2, last_row + 1):
        keszletvaltozas_termek = OrderedDict()
        keszletvaltozas_termek['trafikcikkod'] = sheet.cell(row=x, column=3).value
        keszletvaltozas_termek['megnevezes'] = sheet.cell(row=x, column=5).value
        keszletvaltozas_termek['mennyiseg'] = sheet.cell(row=x, column=7).value
        keszletvaltozas.append(keszletvaltozas_termek)

    j = json.dumps(keszletvaltozas)  # lista json file-ba
    datumkezdet = date.today()
    url = 'http://cleaneffect.hu/ODBE_keszletvaltozasment.php'
    thisdict = dict(datumkezdet = "2023-02-20", megjegyzes = "1 hét", json = j)

    x = requests.post(url, data=thisdict)
    print(x.text)

def odbe_excel_import():
    path = '/Users/futurex/Desktop/2023-02-26_ures_megrendelolap.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    last_row = sheet.max_row
    ODBELista = []

    for i in range(2, last_row + 1):
        ODBETermek = OrderedDict()
        ODBETermek['ndn'] = sheet.cell(row=i, column=1).value
        ODBETermek['megnevezes'] = (sheet.cell(row=i, column=2).value).replace("'","")
        ODBETermek['gyarto'] = sheet.cell(row=i, column=4).value
        ODBETermek['kategoria'] = sheet.cell(row=i, column=5).value
        ODBELista.append(ODBETermek)

    j = json.dumps(ODBELista) # lista json file-ba

    url = 'http://cleaneffect.hu/ODBE_ment.php'
    r = requests.post(url, data=j)
    print(r.text)

def generate_html(dataframe: pd.DataFrame):
    # get the table HTML from the dataframe
    table_html = dataframe.to_html(table_id="table")
    # construct the complete HTML with jQuery Data tables
    # You can disable paging or enable y scrolling on lines 20 and 21 respectively
    html = f"""
    <html>
    <header>
        <link href="https://cdn.datatables.net/1.11.5/css/jquery.dataTables.min.css" rel="stylesheet">
    </header>
    <body>
    {table_html}
    <script src="https://code.jquery.com/jquery-3.6.0.slim.min.js" integrity="sha256-u7e5khyithlIdTpu22PHhENmPcRdFiHRjhAuHcs05RI=" crossorigin="anonymous"></script>
    <script type="text/javascript" src="https://cdn.datatables.net/1.11.5/js/jquery.dataTables.min.js"></script>
    <script>
        $(document).ready( function () {{
            $('#table').DataTable({{
                // paging: false,    
                // scrollY: 400,
            }});
        }});
    </script>
    </body>
    </html>
    """
    # return the html
    return html

